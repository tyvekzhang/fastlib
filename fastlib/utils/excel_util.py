# SPDX-License-Identifier: MIT

"""Excel export utilities for Pydantic models."""

import io
from contextlib import contextmanager
from datetime import datetime
from typing import  Optional, Type

import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from starlette.responses import StreamingResponse


class ExcelExporter:
    """Excel export handler for Pydantic models with customizable styling."""

    DEFAULT_FONT = Font(name="Microsoft YaHei", size=11)
    HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
    HEADER_FILL = PatternFill(
        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
    )
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def __init__(
        self,
        schema: Type[BaseModel],
        file_name: str,
        sheet_name: Optional[str] = None,
        include_timestamp: bool = True,
    ):
        self.schema = schema
        self.file_name = file_name
        self.sheet_name = sheet_name or file_name
        self.include_timestamp = include_timestamp
        self.field_names = list(schema.model_fields.keys())

    def _generate_filename(self) -> str:
        """Generate filename with optional timestamp."""
        if self.include_timestamp:
            timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
            return f"{self.file_name}_{timestamp}.xlsx"
        return f"{self.file_name}.xlsx"

    def _prepare_dataframe(
        self, data_list: Optional[list[BaseModel]] = None
    ) -> pd.DataFrame:
        """Convert Pydantic models to DataFrame."""
        if not data_list:
            return pd.DataFrame(columns=self.field_names)

        data_dicts = [item.model_dump() for item in data_list]
        return pd.DataFrame(data_dicts, columns=self.field_names)

    def _apply_styles(self, worksheet: Worksheet, has_data: bool = False) -> None:
        """Apply consistent styling to the worksheet."""
        # Style header row
        for cell in worksheet[1]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT

        # Style data rows
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = self.DEFAULT_FONT
                cell.alignment = Alignment(vertical="center")

        # Auto-adjust column widths
        for idx, column in enumerate(worksheet.columns, 1):
            column_letter = get_column_letter(idx)
            max_length = 0

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Set minimum width of 12, maximum of 50
            adjusted_width = min(max(max_length + 2, 12), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @contextmanager
    def _excel_writer(self, stream: io.BytesIO):
        """Context manager for Excel writer with error handling."""
        writer = pd.ExcelWriter(stream, engine="openpyxl")
        try:
            yield writer
        finally:
            writer.close()

    async def export(
        self, data_list: Optional[list[BaseModel]] = None
    ) -> StreamingResponse:
        """Export data to Excel file."""
        filename = self._generate_filename()
        stream = io.BytesIO()

        try:
            df = self._prepare_dataframe(data_list)

            with self._excel_writer(stream) as writer:
                df.to_excel(writer, index=False, sheet_name=self.sheet_name)
                worksheet = writer.sheets[self.sheet_name]
                self._apply_styles(worksheet, has_data=bool(data_list))

            stream.seek(0)

            return StreamingResponse(
                stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )

        except Exception as e:
            logger.error(f"Failed to export Excel: {e}")
            raise
